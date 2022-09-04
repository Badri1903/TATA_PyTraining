#                                   ......:::::: PYTHON IN EXCEL :::::......

from openpyxl import Workbook
from datetime import datetime

wb = Workbook()
ws = wb.active

ws.title = 'Data Entry Sheet'

ws['C5'] = 'Hello world'
ws['D5'] = 45250
ws['E5'] = datetime.now()

wb.save('FirstExcel.xlsx')

