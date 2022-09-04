from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

wb.active = wb["Employee"]

ws = wb.active

print(ws.dimensions)

for row in ws.iter_rows (min_row= 5, min_col= 1, max_row= 12, max_col= 5):
    for cell in row:
        print("{:^10s}".format(str(cell.value)), end= " ")
    print()