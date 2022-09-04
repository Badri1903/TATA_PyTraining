from openpyxl import load_workbook

wb = load_workbook('FirstExcel.xlsx')
wb.active = wb['Employee']
ws = wb.active

for row in ws.iter_rows(min_row=5, max_row=12, min_col=2, max_col=2):
    for cell in row:
        if cell.value == 'John':
            cell.value = 'John trovolts'.upper()

wb.save('FirstExcel.xlsx')