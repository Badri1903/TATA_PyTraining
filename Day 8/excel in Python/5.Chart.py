from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart

wb = load_workbook('FirstExcel.xlsx')
wb.create_sheet('MyChart')
wb.active = wb['MyChart']
ws = wb.active

data =[
    ('Product', 'Sales'),
    ('Pepsi', 245),
    ('Coke', 220),
    ('Miranda', 150),
    ('Slice', 200)
]

for i in data:
    ws.append(i)

print(ws.dimensions)

data_ref = Reference(ws, min_col=2, max_col=2, min_row=2, max_row=5)
label_ref = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=5)

chart = BarChart()
chart.add_data(data_ref)
chart.set_categories(label_ref)
chart.title = 'Beverage sales'
chart.x_axis.title = 'Products'
chart.y_axis.title = 'Sales in crates'

ws.add_chart(chart, "E7")
wb.save('FirstExcel.xlsx')


